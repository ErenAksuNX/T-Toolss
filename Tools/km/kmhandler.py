import datetime
import os
from datetime import *
import pandas as pd
from Tools.btExceptions import *
from tkinter.messagebox import showinfo
from openpyxl import load_workbook
from shutil import copyfile
from Tools.HelpFunc import nullstring, removeDupes, dateFormate_dd_mm_YYYY_HHMMSS, getIndex_km


class Km:
    def __init__(self, path, destination):
        self.ws = None
        self.workbook = None
        self.df_auszug = None
        self.df_kmstand = None
        self.path = path
        self.destination = destination
        self.Fahrzeuge = list()

        self.now = datetime.now()

        self.type = ""

        now = self.now

        self.filepath = self.destination + "/KM_Auswertung" + str(int(now.year)) + nullstring(now.month) + nullstring(
            now.day) + ".xlsx"

        self.kmstand = r"O:\00_ECM2\04_Kilometer\Tool\Kilometerstand.xlsx"

        self.filepath = self.destination + "/KM_Auswertung" + str(int(now.year)) + nullstring(now.month) + nullstring(
            now.day) + ".xlsx"

    def run(self):

        # In diesem Block wird überprüft ob der Benutzer den VPN benutzt oder nicht

        try:
            self.df_kmstand = pd.read_excel(self.kmstand)
        except FileNotFoundError as err:
            vpn_error(err)
            return

        # In diesem Block wird gechecked ob eine Excel Tabelle offen ist

        try:
            self.df_kmstand = pd.read_excel(self.kmstand)
            self.df_auszug = pd.read_excel(self.path)
        except PermissionError as err:
            doc_open(err)
            return

        # Hier wird überprüft um welche art vom Tabelle es sich Handelt

        if "Dienste_" in str(self.path):
            self.type = "Dienst"
            self.dienstlist()
        elif "Kilometerstaende_" in self.path:
            self.type = "KM"
            self.kmlist()
        else:
            showinfo(Title="Datei hat einen ungültigen namen",
                     message="Error! Datei hat einen ungültigen Namen! \nDienste Tabelle Beginnt min \"Dienste_\" \nDie Kilometerliste beginnt mit \"Kilometerstaende_\"")

    def dienstlist(self):
        for i, row in self.df_auszug.iterrows():  # Für jede Zeile aus dem Railcloud auszug wird überprüft ob der wert

            index = getIndex_km(self.df_auszug.loc[i, 'Fahrzeug'])

            # In den unteren if Anweisungen wird überprüft, ob die werte in der Tabelle Plausible sind

            if self.df_auszug.loc[i, "Einheit"] != "km":
                continue

            if int(self.df_auszug.at[i, "Wert"]) <= int(self.df_kmstand.at[index - 2, "Kilometerstand"]):
                continue

            if int(self.df_auszug.at[i, "Wert"]) > 10000000:
                showinfo(title="Auffälligkeit Entdeckt", message="Bei dem Fahrzeug " + str(self.df_auszug.at[
                                                                                               i, "Fahrzeug"]) + " ist ein extrem hoher Kilometerstand eingetragen deswegen wird dieses Fahrzeug übersprungen")
                continue

            if int(self.df_auszug.at[i, "Wert"]) < 500000:
                showinfo(title="Auffälligkeit Entdeckt", message="Bei dem Fahrzeug " + str(self.df_auszug.at[
                                                                                               i, "Fahrzeug"]) + " ist ein extrem niedriger wert eingetragen, deswegen wird dieses Fahrzeug übersprungen")
                continue

            # Hier werden die Datensätze in der Tabelle und in der Liste eingetragen
            self.df_kmstand.at[index - 2, "Kilometerstand"] = self.df_auszug.at[i, "Wert"]
            self.Fahrzeuge.append(self.df_auszug.at[i, "Fahrzeug"])
            self.df_kmstand.at[index - 2, "Datum"] = self.df_auszug.at[i, "Erstellt am"]

        # Hier wird eingetragen, wer und wann zuletzt die Tabelle eingetragen

        self.df_kmstand.at[2, "   "] = nullstring(self.now.day) + "." + nullstring(self.now.month) + "." + nullstring(
            self.now.year)
        self.df_kmstand.at[6, "                  "] = os.getlogin()

        self.df_kmstand.to_excel(self.kmstand, index=False)
        self.Fahrzeuge = removeDupes(self.Fahrzeuge)

        self.auszugErstellen()

    def kmlist(self):

        for i, row in self.df_auszug.iterrows():  # Für jede Zeile aus dem Railcloud auszug wird überprüft ob der wert

            # In den unteren if Anweisungen wird überprüft, ob die werte in der Tabelle Plausible sind

            index = getIndex_km(self.df_auszug.loc[i, ' Schienenfahrzeugnummer'])

            if self.df_auszug.at[i, "Kilometerstand"] <= self.df_kmstand.at[index - 2, "Kilometerstand"]:
                continue

            if self.df_auszug.at[i, "Kilometerstand"] > 10000000:
                showinfo(title="Auffälligkeit Entdeckt",
                         message="Bei dem Fahrzeug " + str() + " ist ein extrem hoher Kilometerstand eingetragen deswegen wird dieses Fahrzeug übersprungen")
                continue

            if int(self.df_auszug.at[i, "Kilometerstand"]) < 500000:
                showinfo(title="Auffälligkeit Entdeckt", message="Bei dem Fahrzeug " + str(self.df_auszug.at[
                                                                                               i, "Fahrzeug"]) + " ist ein extrem niedriger wert eingetragen, deswegen wird dieses Fahrzeug übersprungen")

            # Hier werden die Datensätze in der Tabelle und in der Liste eingetragen
            self.Fahrzeuge.append(self.df_auszug.at[i, " Schienenfahrzeugnummer"])
            self.df_kmstand.at[index - 2, "Kilometerstand"] = self.df_auszug.at[i, "Kilometerstand"]
            self.df_kmstand.at[index - 2, "Datum"] = dateFormate_dd_mm_YYYY_HHMMSS(self.getDate(i))

        # Hier wird eingetragen, wer und wann zuletzt die Tabelle eingetragen

        self.df_kmstand.at[2, "   "] = nullstring(self.now.day) + "." + nullstring(self.now.month) + "." + nullstring(
            self.now.year)

        self.df_kmstand.at[6, "                  "] = os.getlogin()

        self.Fahrzeuge = removeDupes(self.Fahrzeuge)

        self.df_kmstand.to_excel(self.kmstand, index=False)

        self.auszugErstellen()

    def auszugErstellen(self):

        try:
            copyfile("pic/Km Blank.xlsx", self.filepath)
        except FileExistsError:
            exists(self.filepath)
            return

        self.workbook = load_workbook(self.filepath)
        self.ws = self.workbook.active

        j = 2

        for i in self.Fahrzeuge:
            index = getIndex_km(i) - 2

            self.ws.cell(j, 1).value = i
            self.ws.cell(j, 2).value = self.df_kmstand.at[index, "Kilometerstand"]
            self.ws.cell(j, 4).value = self.df_kmstand.at[index, "Datum"]

            j = j + 1

        # if self.type == "KM":
        #
        #     for i, row in self.df_auszug.iterrows():
        #         self.ws.insert_rows(i + 2)
        #         self.ws.cell(i + 2, 1).value = self.df_auszug.at[i, " Schienenfahrzeugnummer"]
        #         self.ws.cell(i + 2, 2).value = self.df_auszug.at[i, "Kilometerstand"]
        #         self.ws.cell(i + 2, 4).value = str(self.getDate(i))
        #
        # if self.type == "Dienst":
        #
        #     j = 2
        #     for i, row in self.df_auszug.iterrows():
        #
        #         if self.df_auszug.at[i, "Einheit"] == "km" and self.df_auszug.at[i, "Fahrzeug"] in self.Fahrzeuge:
        #             self.ws.cell(j, 1).value = self.df_auszug.at[i, "Fahrzeug"]
        #             self.ws.cell(j, 2).value = self.df_auszug.at[i, "Wert"]
        #             self.ws.cell(j, 4).value = self.df_auszug.at[i, "Erstellt am"]
        #             j = j + 1

        self.workbook.save(self.filepath)
        showinfo(title="Erfolg!", message="Die Datei wurde erfolgreich abgespeichert")

        # copyfile("pic/km.xlsx", self.filepath)
        #
        # self.workbook = load_workbook(self.filepath)
        # self.worksheet = self.workbook.active
        #
        # for row in range(2, 37):
        #     self.worksheet["B" + str(row)] = self.df_kmstand.at[row - 2, "Kilometerstand"]
        #
        # if self.type == "Dienst":
        #     for row in range(2, 37):
        #         self.worksheet["D" + str(row)] = self.getDate(row)
        #
        # self.workbook.save(self.filepath)
        #
        #

    def getDate(self, i):
        now = datetime.now()
        try:
            dateTable = datetime.strptime(str(self.df_auszug.at[i, "Datum"]), "%d.%m.%Y %H:%M")
        except ValueError:
            dateTable = datetime.strptime(str(self.df_auszug.at[i, "Datum"]), "%Y-%m-%d %H:%M:%S")
        dateTable = dateTable + timedelta(days=1)

        dateday = datetime.strptime(
            nullstring(dateTable.day) + "." + nullstring(dateTable.month) + "." + str(dateTable.year), "%d.%m.%Y")

        if dateday > now:
            return now
        else:
            try:
                dateTable = datetime.strptime(str(self.df_auszug.at[i, "Datum"]), "%d.%m.%Y %H:%M")
            except ValueError:
                dateTable = datetime.strptime(str(self.df_auszug.at[i, "Datum"]), "%Y-%m-%d %H:%M:%S")
            dateTable = dateTable + timedelta(days=-1)
            return datetime.strptime(
                nullstring(dateTable.day) + "." + nullstring(dateTable.month) + "." + str(dateTable.year) + " 23:59",
                "%d.%m.%Y %H:%M")
