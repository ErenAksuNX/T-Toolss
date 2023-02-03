import arrow
from tkinter.messagebox import *
from shutil import copyfile
import openpyxl as xl


class FristHandler:
    def __init__(self, datei, speicherort):
        self.standardTabelle = "pic/f.xlsm"
        self.path = datei
        self.destination = speicherort

    def run(self):
        # in dem Try Catch Block wird versucht die Datei zu erstellen
        try:
            copyfile(self.standardTabelle,
                     self.destination + "/IH-Tool_" + str(arrow.now().format("DD.MM.YYYY")) + ".xlsm")
        except PermissionError:
            showerror(title="Error!!!", message="Eine Datei ist offen !")
            return

        except:
            showerror(title="Fehler !!!", message="Datei konnte nicht erstellt werden !")
            return

        # Hier werden die Path Variablen erstellt
        # Path1 = IH-Vorschläge
        # Path2 = Die Auswertung

        path1 = self.path
        path2 = self.destination + "/IH-Tool_" + str(arrow.now().format("DD.MM.YYYY")) + ".xlsm"

        # Ab hier wird das IH-Vorschläge ws rüberkopiert

        wb1 = xl.load_workbook(filename=path1)
        ws1 = wb1.worksheets[0]

        wb2 = xl.load_workbook(filename=path2, keep_vba=True, read_only=False)
        ws2 = wb2.create_sheet(ws1.title)

        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value

        wb2.save(path2)

        showinfo(title="Erfolg!!!", message="Die Datei wurde erstellt")
